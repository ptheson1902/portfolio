"""XLSX Export service for generating skill sheet from database."""
import os
import re
import zipfile
from io import BytesIO
from typing import Dict, Any, List, Tuple
from dataclasses import dataclass, field
import xml.etree.ElementTree as ET

from sqlalchemy.orm import Session

from ..models.schemas import Language
from ..repositories.profile_repository import ProfileRepository
from ..repositories.skill_repository import SkillRepository
from ..repositories.project_repository import ProjectRepository


@dataclass
class CellData:
    """Represents a single cell in the Excel sheet."""
    value: str = ""
    row: int = 0
    col: int = 0
    rowspan: int = 1
    colspan: int = 1
    is_merged: bool = False
    is_merged_hidden: bool = False
    width: float = 64.0
    height: float = 20.0


@dataclass
class SheetData:
    """Represents a single Excel sheet."""
    name: str
    cells: List[List[CellData]] = field(default_factory=list)
    col_widths: List[float] = field(default_factory=list)
    row_heights: List[float] = field(default_factory=list)
    max_row: int = 0
    max_col: int = 0


class XlsxExportService:
    """Service for exporting skill sheet to XLSX/HTML/PDF formats."""

    LEVEL_COLUMNS = {
        1: "L",
        2: "M",
        3: "N",
        4: "O",
        5: "P",
    }

    def __init__(self, db: Session):
        self.db = db
        self.profile_repo = ProfileRepository(db)
        self.skill_repo = SkillRepository(db)
        self.project_repo = ProjectRepository(db)

        self.template_path = os.path.join(
            os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(__file__)))),
            "Skill_Sheet_Pham_The_Son_ja.xlsx"
        )

    def _fix_xlsx_for_openpyxl(self, xlsx_bytes: bytes) -> bytes:
        """Fix xlsx file to work with openpyxl."""
        output = BytesIO()

        with zipfile.ZipFile(BytesIO(xlsx_bytes), 'r') as zin:
            with zipfile.ZipFile(output, 'w', zipfile.ZIP_DEFLATED) as zout:
                for item in zin.infolist():
                    data = zin.read(item.filename)

                    if item.filename.startswith('xl/worksheets/') and item.filename.endswith('.xml'):
                        try:
                            content = data.decode('utf-8')
                            content = re.sub(r'left=""', 'left="0.7"', content)
                            content = re.sub(r'right=""', 'right="0.7"', content)
                            content = re.sub(r'top=""', 'top="0.75"', content)
                            content = re.sub(r'bottom=""', 'bottom="0.75"', content)
                            content = re.sub(r'header=""', 'header="0.3"', content)
                            content = re.sub(r'footer=""', 'footer="0.3"', content)
                            content = re.sub(r'paperSize=""', 'paperSize="9"', content)
                            content = re.sub(r'scale=""', 'scale="100"', content)
                            content = re.sub(r'orientation=""', 'orientation="portrait"', content)
                            data = content.encode('utf-8')
                        except Exception:
                            pass

                    zout.writestr(item, data)

        output.seek(0)
        return output.getvalue()

    def _gather_data(self, lang: Language) -> Dict[str, Any]:
        """Gather all data needed for the skill sheet."""
        profile = self.profile_repo.get_profile()
        if not profile:
            raise ValueError("Profile not found")

        skills_raw = self.skill_repo.get_skills_grouped_by_category(lang, None)
        projects = self.project_repo.get_projects_for_resume(lang)

        return {
            "profile": profile,
            "skills": skills_raw,
            "projects": projects,
        }

    def _format_experience(self, experience: str) -> str:
        """Format experience string."""
        if not experience:
            return "0 年 - 0 ヶ月"

        years = 0
        months = 0

        year_match = re.search(r'(\d+)\s*年', experience)
        month_match = re.search(r'(\d+)\s*ヶ月', experience)

        if year_match:
            years = int(year_match.group(1))
        if month_match:
            months = int(month_match.group(1))

        return f"{years} 年 - {months} ヶ月"

    def _col_letter_to_index(self, col: str) -> int:
        """Convert column letter to index (A=0, B=1, etc)."""
        result = 0
        for char in col:
            result = result * 26 + (ord(char.upper()) - ord('A') + 1)
        return result - 1

    def _parse_cell_ref(self, ref: str) -> Tuple[int, int]:
        """Parse cell reference like 'A1' to (row, col)."""
        match = re.match(r'([A-Z]+)(\d+)', ref)
        if match:
            col = self._col_letter_to_index(match.group(1))
            row = int(match.group(2)) - 1
            return row, col
        return 0, 0

    def _load_shared_strings(self, zf: zipfile.ZipFile) -> List[str]:
        """Load shared strings from xlsx."""
        strings = []
        try:
            with zf.open('xl/sharedStrings.xml') as f:
                tree = ET.parse(f)
                root = tree.getroot()
                ns = {'main': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
                for si in root.findall('.//main:si', ns):
                    text_parts = []
                    for t in si.findall('.//main:t', ns):
                        if t.text:
                            text_parts.append(t.text)
                    strings.append(''.join(text_parts))
        except KeyError:
            pass
        return strings

    def _fix_sheet_xml(self, xml_bytes: bytes) -> bytes:
        """Fix problematic empty attributes in sheet XML."""
        try:
            content = xml_bytes.decode('utf-8')
            content = re.sub(r'left=""', 'left="0.7"', content)
            content = re.sub(r'right=""', 'right="0.7"', content)
            content = re.sub(r'top=""', 'top="0.75"', content)
            content = re.sub(r'bottom=""', 'bottom="0.75"', content)
            content = re.sub(r'header=""', 'header="0.3"', content)
            content = re.sub(r'footer=""', 'footer="0.3"', content)
            content = re.sub(r'paperSize=""', 'paperSize="9"', content)
            content = re.sub(r'scale=""', 'scale="100"', content)
            content = re.sub(r'orientation=""', 'orientation="portrait"', content)
            return content.encode('utf-8')
        except Exception:
            return xml_bytes

    def _parse_xlsx_to_sheets(self, xlsx_bytes: bytes) -> List[SheetData]:
        """Parse xlsx file to sheet data structures."""
        sheets = []
        ns = {'main': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}

        with zipfile.ZipFile(BytesIO(xlsx_bytes), 'r') as zf:
            shared_strings = self._load_shared_strings(zf)

            # Get sheet names
            with zf.open('xl/workbook.xml') as f:
                tree = ET.parse(f)
                root = tree.getroot()
                sheets_elem = root.findall('.//main:sheet', ns)
                sheet_names = [s.get('name', f'Sheet{i+1}') for i, s in enumerate(sheets_elem)]

            # Load each sheet
            for i, name in enumerate(sheet_names):
                sheet_path = f'xl/worksheets/sheet{i+1}.xml'
                try:
                    with zf.open(sheet_path) as f:
                        xml_content = self._fix_sheet_xml(f.read())
                        tree = ET.parse(BytesIO(xml_content))
                        root = tree.getroot()

                    sheet = SheetData(name=name)

                    # Get column widths
                    cols = root.findall('.//main:col', ns)
                    col_widths = {}
                    for col in cols:
                        min_col = int(col.get('min', 1)) - 1
                        max_col = int(col.get('max', 1)) - 1
                        width = float(col.get('width', 8.43)) * 7
                        for c in range(min_col, max_col + 1):
                            col_widths[c] = width

                    # Get row heights
                    rows = root.findall('.//main:row', ns)
                    row_heights = {}
                    for row in rows:
                        row_num = int(row.get('r', 1)) - 1
                        height = float(row.get('ht', 15))
                        row_heights[row_num] = height

                    # Get merged cells
                    merged_cells = {}
                    merge_cells = root.findall('.//main:mergeCell', ns)
                    for merge in merge_cells:
                        ref = merge.get('ref', '')
                        if ':' in ref:
                            start, end = ref.split(':')
                            start_row, start_col = self._parse_cell_ref(start)
                            end_row, end_col = self._parse_cell_ref(end)
                            rowspan = end_row - start_row + 1
                            colspan = end_col - start_col + 1
                            merged_cells[(start_row, start_col)] = (rowspan, colspan)
                            for r in range(start_row, end_row + 1):
                                for c in range(start_col, end_col + 1):
                                    if r != start_row or c != start_col:
                                        merged_cells[(r, c)] = (0, 0)

                    max_row = max(row_heights.keys(), default=0) + 1
                    max_col = max(col_widths.keys(), default=0) + 1
                    max_row = max(max_row, 60)
                    max_col = max(max_col, 20)

                    sheet.col_widths = [col_widths.get(c, 64) for c in range(max_col)]
                    sheet.row_heights = [row_heights.get(r, 20) for r in range(max_row)]
                    sheet.max_row = max_row
                    sheet.max_col = max_col

                    # Create empty grid
                    sheet.cells = []
                    for r in range(max_row):
                        row_cells = []
                        for c in range(max_col):
                            cell = CellData(row=r, col=c)
                            cell.width = sheet.col_widths[c] if c < len(sheet.col_widths) else 64
                            cell.height = sheet.row_heights[r] if r < len(sheet.row_heights) else 20

                            if (r, c) in merged_cells:
                                rowspan, colspan = merged_cells[(r, c)]
                                if rowspan == 0 and colspan == 0:
                                    cell.is_merged_hidden = True
                                else:
                                    cell.is_merged = True
                                    cell.rowspan = rowspan
                                    cell.colspan = colspan

                            row_cells.append(cell)
                        sheet.cells.append(row_cells)

                    # Fill cell values
                    for row in rows:
                        for c in row.findall('main:c', ns):
                            ref = c.get('r', '')
                            cell_row, cell_col = self._parse_cell_ref(ref)
                            cell_type = c.get('t', '')

                            v = c.find('main:v', ns)
                            value = ""
                            if v is not None and v.text:
                                if cell_type == 's':
                                    idx = int(v.text)
                                    if idx < len(shared_strings):
                                        value = shared_strings[idx]
                                else:
                                    value = v.text

                            is_elem = c.find('main:is', ns)
                            if is_elem is not None:
                                t_elem = is_elem.find('main:t', ns)
                                if t_elem is not None and t_elem.text:
                                    value = t_elem.text

                            if cell_row < len(sheet.cells) and cell_col < len(sheet.cells[cell_row]):
                                sheet.cells[cell_row][cell_col].value = value

                    sheets.append(sheet)
                except KeyError:
                    continue

        return sheets

    def generate_xlsx(self, lang: Language) -> bytes:
        """Generate XLSX skill sheet by modifying template."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise ImportError("openpyxl is required. Install with: pip install openpyxl")

        if not os.path.exists(self.template_path):
            raise FileNotFoundError(f"Template not found: {self.template_path}")

        with open(self.template_path, 'rb') as f:
            xlsx_bytes = f.read()

        fixed_xlsx = self._fix_xlsx_for_openpyxl(xlsx_bytes)
        wb = load_workbook(BytesIO(fixed_xlsx))
        ws = wb.active

        data = self._gather_data(lang)
        profile = data["profile"]
        skills = data["skills"]
        projects = data["projects"]

        # Update profile section
        ws['C4'] = profile.name_vi if lang == Language.VI else profile.name
        ws['F4'] = profile.name_kana
        ws['C5'] = f"{profile.school}({profile.graduation_year}年卒業)"
        ws['H5'] = profile.age
        ws['C6'] = profile.field
        ws['C7'] = profile.work_experience
        ws['F7'] = profile.japan_residence
        ws['H7'] = profile.japanese_level

        self_pr = profile.self_pr.get(lang.value, profile.self_pr.get("ja", "")) if profile.self_pr else ""
        ws['C9'] = self_pr  # Full self PR

        # Update skills section
        category_mapping = {
            "programming_languages": list(range(11, 17)),
            "frameworks": list(range(18, 25)),
            "cloud": [26],
            "databases": list(range(28, 33)),
            "ai_ml": [34],
        }

        for cat_key, skill_list in skills.items():
            if cat_key in category_mapping:
                rows = category_mapping[cat_key]
                for i, skill in enumerate(skill_list):
                    if i >= len(rows):
                        break
                    row = rows[i]
                    ws[f'K{row}'] = skill.get('name', '')
                    level = skill.get('level', 3)
                    if 1 <= level <= 5:
                        level_col = self.LEVEL_COLUMNS.get(level, 'N')
                        ws[f'{level_col}{row}'] = ' ● '
                    exp = skill.get('experience', '')
                    ws[f'Q{row}'] = self._format_experience(exp)

        # Update projects section
        project_rows = [14, 18, 22, 26, 30, 34, 38, 42, 46, 50]

        for i, project in enumerate(projects):
            if i >= len(project_rows):
                break

            row = project_rows[i]
            ws[f'A{row}'] = i + 1
            ws[f'B{row}'] = project.get('name', '')

            if row + 1 <= ws.max_row:
                ws[f'B{row + 1}'] = project.get('description', '')  # Full description

            ws[f'D{row}'] = project.get('role', '')
            if row + 1 <= ws.max_row:
                ws[f'D{row + 1}'] = f"チーム: {project.get('team_size', 0)} 名"

            techs = project.get('technologies', [])
            ws[f'E{row}'] = '\n'.join(techs[:3]) if techs else ""
            ws[f'F{row}'] = '\n'.join(techs[3:6]) if len(techs) > 3 else ""
            ws[f'H{row}'] = project.get('start_date', '')

            end_date = project.get('end_date', '')
            if end_date and row + 2 <= ws.max_row:
                ws[f'H{row + 2}'] = end_date

            ws[f'I{row}'] = project.get('duration', '')

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        wb.close()

        return output.getvalue()

    def generate_html(self, lang: Language) -> str:
        """Generate HTML skill sheet matching Excel layout."""
        data = self._gather_data(lang)
        profile = data["profile"]
        skills = data["skills"]
        projects = data["projects"]

        # Get self_pr with proper language
        self_pr = ""
        if profile.self_pr:
            self_pr = profile.self_pr.get(lang.value, profile.self_pr.get("ja", ""))

        # Build skills by category
        skill_categories = {
            "programming_languages": {"label": "プログラミング言語", "skills": []},
            "frameworks": {"label": "フレームワーク/ IDE", "skills": []},
            "cloud": {"label": "クラウドサービス", "skills": []},
            "databases": {"label": "データベース", "skills": []},
            "ai_ml": {"label": "ドメイン", "skills": []},
        }

        for cat_key, skill_list in skills.items():
            if cat_key in skill_categories:
                skill_categories[cat_key]["skills"] = skill_list

        return self._render_html_template(profile, skill_categories, projects, self_pr, lang)

    def _render_html_template(self, profile, skill_categories, projects, self_pr, lang) -> str:
        """Render HTML template with data."""

        def level_dots(level: int) -> str:
            """Generate level indicator."""
            cols = ["", "", "", "", ""]
            if 1 <= level <= 5:
                cols[level - 1] = "●"
            return "".join([f'<td class="level-cell">{c}</td>' for c in cols])

        html = '''<!DOCTYPE html>
<html lang="ja">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>情報処理技術者経歴書</title>
    <style>
        @import url('https://fonts.googleapis.com/css2?family=Noto+Sans+JP:wght@400;500;700&display=swap');

        * { margin: 0; padding: 0; box-sizing: border-box; }
        html { -webkit-print-color-adjust: exact !important; print-color-adjust: exact !important; }

        body {
            font-family: 'Noto Sans JP', 'Hiragino Kaku Gothic ProN', 'Meiryo', sans-serif;
            font-size: 10px;
            line-height: 1.4;
            background: #f0f0f0;
            color: #222;
        }

        .page {
            width: 297mm;
            min-height: 210mm;
            margin: 10mm auto;
            background: white;
            padding: 8mm;
            box-shadow: 0 2px 10px rgba(0,0,0,0.15);
        }

        h1.title {
            text-align: center;
            font-size: 18px;
            font-weight: 700;
            letter-spacing: 0.8em;
            margin-bottom: 6mm;
            padding-bottom: 3mm;
            border-bottom: 2px solid #333;
        }

        .main-layout {
            display: flex;
            gap: 4mm;
        }

        .left-section { flex: 6; }
        .right-section { flex: 4; min-width: 100mm; }

        table { border-collapse: collapse; width: 100%; font-size: 9px; }
        th, td {
            border: 1px solid #999;
            padding: 2px 4px;
            vertical-align: top;
            text-align: left;
        }

        .label { background: #e8e8e8; font-weight: 500; text-align: center; width: 80px; }
        .section-header { background: #d0d8e8; font-weight: 600; text-align: center; }
        .category-header { background: #e3f0ff; font-weight: 600; }

        .profile-table td { height: 22px; }
        .profile-table .value { min-width: 120px; }

        .pr-cell {
            height: 60px;
            white-space: pre-wrap;
            word-wrap: break-word;
            font-size: 8px;
            line-height: 1.5;
        }

        .skill-table { margin-top: 2mm; }
        .skill-table th { text-align: center; font-size: 8px; }
        .skill-name { width: 90px; }
        .level-cell { width: 20px; text-align: center; font-weight: bold; }
        .exp-cell { width: 55px; text-align: center; font-size: 8px; }

        .project-table { margin-top: 3mm; font-size: 8px; }
        .project-table th { background: #e8e8e8; text-align: center; font-size: 8px; padding: 3px 2px; }
        .project-table td { padding: 2px 3px; vertical-align: top; }
        .project-num { width: 20px; text-align: center; font-weight: bold; }
        .project-content { width: auto; white-space: pre-wrap; word-wrap: break-word; }
        .project-role { width: 55px; }
        .project-tech { width: 70px; white-space: pre-wrap; font-size: 7px; }
        .project-env { width: 65px; white-space: pre-wrap; font-size: 7px; }
        .project-phase { width: 55px; white-space: pre-wrap; font-size: 7px; }
        .project-period { width: 50px; text-align: center; font-size: 7px; }
        .project-duration { width: 40px; text-align: center; }

        .ranking-legend { margin-top: 2mm; font-size: 7px; }
        .ranking-legend td { padding: 1px 3px; }

        @media print {
            body { background: white; }
            .page { margin: 0; padding: 5mm; box-shadow: none; page-break-after: always; }
        }

        @page { size: A4 landscape; margin: 5mm; }
    </style>
</head>
<body>
<div class="page">
    <h1 class="title">情 報 処 理 技 術 者 経 歴 書</h1>

    <div class="main-layout">
        <!-- LEFT SECTION: Profile + Projects -->
        <div class="left-section">
            <!-- Profile Info -->
            <table class="profile-table">
                <tr>
                    <td class="label">氏名</td>
                    <td class="value" colspan="2">''' + (profile.name_vi if lang == Language.VI else profile.name) + '''</td>
                    <td class="label">カナ</td>
                    <td class="value" colspan="2">''' + (profile.name_kana or "") + '''</td>
                    <td class="label">性別</td>
                    <td>男</td>
                </tr>
                <tr>
                    <td class="label">出身校</td>
                    <td class="value" colspan="4">''' + f"{profile.school}（{profile.graduation_year}年卒業）" + '''</td>
                    <td class="label">年齢</td>
                    <td colspan="2">''' + str(profile.age) + '''</td>
                </tr>
                <tr>
                    <td class="label">対応可能分野</td>
                    <td class="value" colspan="4">''' + (profile.field or "") + '''</td>
                    <td class="label">英語レベル</td>
                    <td colspan="2">-</td>
                </tr>
                <tr>
                    <td class="label">業務経験</td>
                    <td class="value">''' + (profile.work_experience or "") + '''</td>
                    <td class="label">日本常駐</td>
                    <td class="value" colspan="2">''' + (profile.japan_residence or "") + '''</td>
                    <td class="label">日本語レベル</td>
                    <td colspan="2">''' + (profile.japanese_level or "") + '''</td>
                </tr>
                <tr>
                    <td class="label">自己PR</td>
                    <td class="pr-cell" colspan="7">''' + self_pr.replace('\n', '<br>') + '''</td>
                </tr>
            </table>

            <!-- Projects -->
            <table class="project-table">
                <tr>
                    <th class="project-num">No</th>
                    <th class="project-content">業務内容</th>
                    <th class="project-role">役割/規模</th>
                    <th class="project-tech">使用言語</th>
                    <th class="project-env">環境</th>
                    <th class="project-phase">担当フェーズ</th>
                    <th class="project-period">開始/終了</th>
                    <th class="project-duration">期間</th>
                </tr>
'''

        # Add projects
        for i, project in enumerate(projects):
            name = project.get('name', '')
            desc = project.get('description', '')
            role = project.get('role', '')
            team = project.get('team_size', 0)
            techs = project.get('technologies', [])
            start = project.get('start_date', '')
            end = project.get('end_date', '')
            duration = project.get('duration', '')
            highlights = project.get('highlights', [])

            # Split technologies
            lang_techs = '\n'.join(techs[:4]) if techs else ''
            env_techs = '\n'.join(techs[4:8]) if len(techs) > 4 else ''

            # Phase info
            phase = "製造\n単体テスト"
            if role and 'リーダー' in role:
                phase = "基本設計\n詳細設計\nコーディング\n単体テスト"
            elif role and 'BrSE' in role:
                phase = "要件定義\n詳細設計\nQ&A対応"

            html += f'''
                <tr>
                    <td class="project-num" rowspan="2">{i+1}</td>
                    <td class="project-content"><b>{name}</b></td>
                    <td class="project-role">{role}</td>
                    <td class="project-tech" rowspan="2">{lang_techs}</td>
                    <td class="project-env" rowspan="2">{env_techs}</td>
                    <td class="project-phase" rowspan="2">{phase}</td>
                    <td class="project-period">{start}</td>
                    <td class="project-duration" rowspan="2">{duration}</td>
                </tr>
                <tr>
                    <td class="project-content" style="font-size:7px;">{desc}</td>
                    <td class="project-role" style="font-size:7px;">チーム: {team}名</td>
                    <td class="project-period" style="font-size:7px;">{end}</td>
                </tr>
'''

        html += '''
            </table>
        </div>

        <!-- RIGHT SECTION: Skills -->
        <div class="right-section">
            <!-- Ranking Legend -->
            <table class="ranking-legend">
                <tr>
                    <td class="section-header" colspan="3">ランキングの説明</td>
                </tr>
                <tr><td>1</td><td colspan="2">基本的な理解があり、指導を受けながら作業できる</td></tr>
                <tr><td>2</td><td colspan="2">独力で標準的な作業を遂行できる</td></tr>
                <tr><td>3</td><td colspan="2">高度な作業も含め、自立して実行できる</td></tr>
                <tr><td>4</td><td colspan="2">他者を指導し、プロジェクトをリードできる</td></tr>
                <tr><td>5</td><td colspan="2">エキスパートレベル、技術選定や設計を主導</td></tr>
            </table>

            <!-- Skills -->
            <table class="skill-table">
                <tr>
                    <th class="section-header" colspan="7">スキル一覧</th>
                </tr>
                <tr>
                    <th></th>
                    <th class="level-cell">1</th>
                    <th class="level-cell">2</th>
                    <th class="level-cell">3</th>
                    <th class="level-cell">4</th>
                    <th class="level-cell">5</th>
                    <th class="exp-cell">経験年数</th>
                </tr>
'''

        # Add skills by category
        for cat_key, cat_data in skill_categories.items():
            if cat_data["skills"]:
                html += f'<tr><td class="category-header" colspan="7">{cat_data["label"]}</td></tr>\n'
                for skill in cat_data["skills"]:
                    name = skill.get('name', '')
                    level = skill.get('level', 3)
                    exp = skill.get('experience', '')
                    html += f'<tr><td class="skill-name">{name}</td>{level_dots(level)}<td class="exp-cell">{exp}</td></tr>\n'

        html += '''
            </table>
        </div>
    </div>
</div>
</body>
</html>'''

        return html

    def generate_pdf(self, lang: Language) -> bytes:
        """Generate PDF skill sheet from HTML."""
        html_content = self.generate_html(lang)

        # Try WeasyPrint first
        try:
            from weasyprint import HTML, CSS
            from weasyprint.text.fonts import FontConfiguration

            font_config = FontConfiguration()
            pdf_css = CSS(string='''
                @page { size: A4 landscape; margin: 8mm; }
                body { font-family: 'Noto Sans JP', sans-serif; }
            ''', font_config=font_config)

            pdf_bytes = HTML(string=html_content).write_pdf(
                stylesheets=[pdf_css],
                font_config=font_config
            )
            return pdf_bytes

        except ImportError:
            pass
        except OSError as e:
            if "libgobject" not in str(e) and "cannot load library" not in str(e):
                raise

        # Try wkhtmltopdf
        try:
            import subprocess
            import tempfile

            with tempfile.NamedTemporaryFile(mode='w', suffix='.html', delete=False, encoding='utf-8') as f:
                f.write(html_content)
                html_path = f.name

            with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as f:
                pdf_path = f.name

            try:
                cmd = [
                    "wkhtmltopdf",
                    "--page-size", "A4",
                    "--orientation", "Landscape",
                    "--margin-top", "8mm",
                    "--margin-bottom", "8mm",
                    "--margin-left", "10mm",
                    "--margin-right", "10mm",
                    "--encoding", "UTF-8",
                    "--enable-local-file-access",
                    html_path,
                    pdf_path
                ]

                result = subprocess.run(cmd, capture_output=True)
                if result.returncode == 0:
                    with open(pdf_path, 'rb') as f:
                        return f.read()
            finally:
                if os.path.exists(html_path):
                    os.unlink(html_path)
                if os.path.exists(pdf_path):
                    os.unlink(pdf_path)

        except FileNotFoundError:
            pass

        # Try Playwright (run in thread to avoid async conflict)
        try:
            import tempfile
            import concurrent.futures

            def run_playwright(html_content: str) -> bytes:
                from playwright.sync_api import sync_playwright

                with tempfile.NamedTemporaryFile(mode='w', suffix='.html', delete=False, encoding='utf-8') as f:
                    f.write(html_content)
                    html_path = f.name

                try:
                    with sync_playwright() as p:
                        browser = p.chromium.launch()
                        page = browser.new_page()
                        page.goto(f"file:///{html_path.replace(os.sep, '/')}")
                        page.wait_for_load_state("networkidle")

                        pdf_bytes = page.pdf(
                            format="A4",
                            landscape=True,
                            margin={"top": "8mm", "bottom": "8mm", "left": "10mm", "right": "10mm"},
                            print_background=True
                        )
                        browser.close()
                        return pdf_bytes
                finally:
                    if os.path.exists(html_path):
                        os.unlink(html_path)

            # Run in a separate thread to avoid asyncio conflict
            with concurrent.futures.ThreadPoolExecutor() as executor:
                future = executor.submit(run_playwright, html_content)
                return future.result(timeout=60)

        except ImportError:
            pass
        except Exception as e:
            if "asyncio" not in str(e).lower():
                raise

        raise ImportError(
            "PDF generation requires one of: WeasyPrint (with GTK), wkhtmltopdf, or Playwright. "
            "Install with: pip install playwright && playwright install chromium"
        )
